"""
Verificação rápida de processos via Jusbrasil.
Usa nodriver (CDP async) para automação do Chrome.

Instalação:
    pip install nodriver openpyxl

Uso:
    python verificar_jusbrasil.py
    python verificar_jusbrasil.py --arquivo "caminho/planilha.xlsx"
    python verificar_jusbrasil.py --inicio 51
    python verificar_jusbrasil.py --debug
"""

import argparse
import asyncio
import logging
import os
import random
import re
import shutil
import sys
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import nodriver as uc
from urllib.parse import quote_plus

# ---------------------------------------------------------------------------
# Configurações
# ---------------------------------------------------------------------------
# Adaptado para funcionar tanto localmente quanto no GitHub Actions
ARQUIVO_PADRAO = "processos.xlsx" if os.environ.get("CI") else r"C:\Users\paollo\Downloads\Codigo\Processos arquivados (todos) (40).xlsx"
PERFIL_CHROME  = str(Path.cwd() / ".chrome_jusbrasil")
URL_JB         = "https://www.jusbrasil.com.br/consulta-processual/"
SEL_BUSCA      = "input[data-testid='search-input']"
TIMEOUT        = 20   # segundos

COL_CNJ       = 2    # B
COL_RESULTADO = 8    # H
LINHA_CAB     = 2
LINHA_INICIO  = 3

TERMOS_BAIXADO = [
    "baixado", "arquivado", "encerrado", "transitado em julgado",
    "extinto", "cumprida a sentença", "liquidado", "arquivamento",
    "cumprimento de sentença encerrado", "processo encerrado",
    "ao arquivo", "arquivado definitivamente"
]

TERMOS_PUBLICACAO = [
    "enviado para publicação", "remetido para publicação",
    "aguardando publicação", "em publicação", "minuta publicada",
    "publicado no djt", "publicado no dj", "publicação realizada",
]

_RE_DATA = re.compile(
    r"(?<!\d)"
    r"((?:0?[1-9]|[12]\d|3[01])[/\-\.](?:0?[1-9]|1[0-2])[/\-\.](?:19|20)\d{2}"
    r"|\d{1,2}\s+de\s+"
    r"(?:janeiro|fevereiro|março|abril|maio|junho|julho|agosto"
    r"|setembro|outubro|novembro|dezembro)"
    r"\s+de\s+(?:19|20)\d{2})"
    r"(?!\d)"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("verificar_jusbrasil.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Modelo
# ---------------------------------------------------------------------------
@dataclass
class Processo:
    linha: int
    cnj: str
    resultado: str = ""


# ---------------------------------------------------------------------------
# Planilha
# ---------------------------------------------------------------------------
def _abrir_planilha(caminho: Path):
    log.info(f"Abrindo: {caminho.name}")
    return openpyxl.load_workbook(str(caminho))


def _salvar_seguro(wb, caminho: Path) -> bool:
    tmp = caminho.with_stem(caminho.stem + "_tmp")
    bak = caminho.with_stem(caminho.stem + "_prev")
    try:
        wb.save(str(tmp))
    except Exception as e:
        log.error(f"Erro ao salvar temporário: {e}")
        tmp.unlink(missing_ok=True)
        return False
    if tmp.stat().st_size < 1000:
        log.error("Arquivo temporário muito pequeno — abortando.")
        tmp.unlink(missing_ok=True)
        return False
    if caminho.exists():
        try:
            shutil.copy2(str(caminho), str(bak))
        except Exception:
            pass
    try:
        shutil.move(str(tmp), str(caminho))
        return True
    except Exception as e:
        log.error(f"Erro ao substituir arquivo: {e}")
        return False


def _ler_processos(ws) -> list[Processo]:
    processos = []
    for row in range(LINHA_INICIO, ws.max_row + 1):
        cnj = ws.cell(row, COL_CNJ).value
        if cnj:
            processos.append(Processo(linha=row, cnj=str(cnj).strip()))
    return processos


def _gravar(ws, p: Processo):
    if not ws.cell(LINHA_CAB, COL_RESULTADO).value:
        ws.cell(LINHA_CAB, COL_RESULTADO).value = "Baixado?"
    ws.cell(p.linha, COL_RESULTADO).value = p.resultado


# ---------------------------------------------------------------------------
# Classificação
# ---------------------------------------------------------------------------
def _classificar(texto: str) -> str:
    t = texto.lower()
    if any(termo in t for termo in TERMOS_BAIXADO):
        return "baixado"
    if any(termo in t for termo in TERMOS_PUBLICACAO):
        return "publicacao"
    return "nao"


def _extrair_data(texto: str) -> str:
    t = texto.lower()
    
    # Procura por datas no formato dd/mm/aaaa em todo o texto
    datas_encontradas = _RE_DATA.findall(t)
    
    for termo in TERMOS_BAIXADO + TERMOS_PUBLICACAO:
        idx = t.find(termo)
        if idx >= 0:
            # Pega um trecho maior ao redor do termo encontrado
            trecho = t[max(0, idx - 300): idx + 300]
            m = _RE_DATA.search(trecho)
            if m:
                return m.group(0)
                
    # Se não achou data perto do termo, retorna a primeira data que aparecer no texto (geralmente a mais recente no topo)
    if datas_encontradas:
        return datas_encontradas[0]
        
    return ""


# ---------------------------------------------------------------------------
# Cloudflare helper
# ---------------------------------------------------------------------------
_SINAIS_CF = [
    "challenges.cloudflare.com",
    "cf-turnstile",
    "cf_challenge",
    "Verificando se você é humano",
    "Checking if the site connection is secure",
    "Just a moment",
    "Enable JavaScript and cookies",
]


async def _tem_cloudflare(tab) -> bool:
    try:
        conteudo = await tab.get_content()
        return any(s in conteudo for s in _SINAIS_CF)
    except Exception:
        return False


async def _resolver_cloudflare(tab) -> bool:
    """
    Detecta Cloudflare Turnstile e tenta resolver.
    - Tenta automático por 8s (o Turnstile às vezes passa sozinho com zendriver)
    - Se não resolver, pausa e pede resolução manual no browser
    """
    if not await _tem_cloudflare(tab):
        return True  # Sem desafio — página ok

    log.info("  Cloudflare detectado — tentando resolver automaticamente...")

    # Aguarda até 8s para o zendriver passar sozinho (funciona às vezes)
    for _ in range(8):
        await asyncio.sleep(1)
        if not await _tem_cloudflare(tab):
            log.info("  Cloudflare resolvido automaticamente.")
            return True

    # Fallback: resolução manual (se não estiver no GitHub Actions)
    if os.environ.get("CI"):
        log.warning("  [CI] Ambiente automatizado detectado. Pulando resolução manual do Cloudflare.")
        return False
        
    log.info("")
    log.info("=" * 55)
    log.info("  CLOUDFLARE — resolução manual necessária")
    log.info("  Olhe o browser, resolva o desafio e pressione Enter:")
    log.info("=" * 55)
    try:
        input()
    except EOFError:
        pass

    # Aguarda mais 3s após o Enter para a página carregar
    await asyncio.sleep(3)

    if await _tem_cloudflare(tab):
        log.warning("  Cloudflare ainda presente após resolução manual.")
        return False

    log.info("  Cloudflare resolvido.")
    return True


async def _url_atual(tab) -> str:
    """Retorna a URL atual da aba via JavaScript."""
    try:
        return await tab.evaluate("location.href") or ""
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Jusbrasil
# ---------------------------------------------------------------------------
async def _pesquisar(browser, tab, cnj: str) -> str:
    """
    Pesquisa um CNJ no Jusbrasil e retorna:
    "Sim | dd/mm/aaaa", "Sim", "Não", "Publicação | ...", "Não encontrado", "Erro"
    """
    try:
        # 1. Vai para a home de consulta processual
        try:
            await asyncio.wait_for(tab.get("https://www.jusbrasil.com.br/consulta-processual/"), timeout=30)
        except Exception:
            pass
        await asyncio.sleep(2)
        
        await _resolver_cloudflare(tab)

        # 2. Digita o CNJ no campo de busca e pressiona Enter
        try:
            # Usa uma injeção JS robusta para preencher o campo React e clicar no botão
            sucesso_js = await tab.evaluate(f"""
                (() => {{
                    const cnj = "{cnj}";
                    const inputs = document.querySelectorAll('input[type="search"], input[name="q"]');
                    
                    for (const input of inputs) {{
                        // Verifica se o input está visível
                        if (input.offsetParent !== null) {{
                            input.focus();
                            
                            // Hack necessário para o React reconhecer a mudança de valor
                            const nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, "value").set;
                            nativeInputValueSetter.call(input, cnj);
                            
                            input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                            input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            
                            // Procura o formulário e o botão de submit
                            const form = input.closest('form');
                            if (form) {{
                                const btn = form.querySelector('button[type="submit"], button[aria-label="pesquisar"]');
                                if (btn) {{
                                    btn.click();
                                    return true;
                                }}
                                form.submit();
                                return true;
                            }}
                            
                            // Se não tiver form, tenta achar um botão próximo
                            const container = input.closest('div');
                            if (container) {{
                                const btn = container.querySelector('button');
                                if (btn) {{
                                    btn.click();
                                    return true;
                                }}
                            }}
                        }}
                    }}
                    return false;
                }})()
            """)
            
            if not sucesso_js:
                log.warning(f"  Não foi possível encontrar o campo de busca visível para o CNJ {cnj}")
                return "Erro"
                
        except Exception as e:
            log.warning(f"  Erro ao tentar buscar: {e}")
            return "Erro"

        # Aguarda a URL mudar (sinal de que a busca começou)
        for _ in range(15):
            url_check = await _url_atual(tab)
            if url_check != "https://www.jusbrasil.com.br/consulta-processual/":
                break
            await asyncio.sleep(1)
            
        log.info(f"  URL após busca: {url_check}")

        await _resolver_cloudflare(tab)

        # 4. Detecta resultado
        cnj_digits = re.sub(r"[^\d]", "", cnj)
        cnj_final_13 = cnj_digits[-13:] if len(cnj_digits) >= 13 else cnj_digits
        href = None
        prazo = time.monotonic() + TIMEOUT
        clicou_consulta = False

        for tentativa in range(1, 30):
            if time.monotonic() > prazo:
                break

            url_atual = await _url_atual(tab)
            log.debug(f"  Tentativa {tentativa}: URL = {url_atual}")

            # Caso: Redirecionou para página de processo
            if "/processos/" in url_atual or "/processo/" in url_atual:
                page_text = await tab.evaluate("document.body.innerText") or ""
                page_text_digits = re.sub(r"[^\d]", "", page_text[:5000])
                url_digits = re.sub(r"[^\d]", "", url_atual)

                if (cnj in page_text or
                    cnj_digits in page_text_digits or
                    cnj_final_13 in page_text_digits or
                    cnj_final_13 in url_digits):
                    href = url_atual
                    break
                else:
                    # Se redirecionou para um processo errado, volta para a busca
                    log.warning(f"  Página não contém CNJ {cnj} — processo errado")
                    return "Não encontrado"

            # Caso: Página de resultados com links
            href_found = await tab.evaluate(f"""
                (() => {{
                    const cnj = "{cnj}";
                    const cnjDigits = "{cnj_digits}";
                    const cnjFinal13 = "{cnj_final_13}";
                    
                    // Primeiro tenta achar links diretos para processos
                    const links = document.querySelectorAll('a[href*="/processos/"], a[href*="/processo/"]');
                    for (const a of links) {{
                        const container = a.closest('li, article, div[class*="result"], div[class*="card"], div[class*="ProcessSnippet"]') || a.parentElement;
                        const text = (container ? container.innerText : a.innerText) || "";
                        const textDigits = text.replace(/[^\\d]/g, "");
                        const hrefDigits = (a.href || "").replace(/[^\\d]/g, "");
                        if (text.includes(cnj) || textDigits.includes(cnjDigits) || textDigits.includes(cnjFinal13) || hrefDigits.includes(cnjFinal13)) {{
                            return a.href || a.getAttribute("href") || "";
                        }}
                    }}
                    
                    // Se não achou nos links, tenta achar em qualquer elemento clicável que tenha o CNJ
                    const allEls = document.querySelectorAll('div, span, p, h1, h2, h3, h4, h5, h6');
                    for (const el of allEls) {{
                        if (el.innerText && (el.innerText.includes(cnj) || el.innerText.replace(/[^\\d]/g, "").includes(cnjDigits))) {{
                            const link = el.closest('a');
                            if (link && (link.href.includes('/processos/') || link.href.includes('/processo/'))) {{
                                return link.href;
                            }}
                        }}
                    }}
                    
                    return "";
                }})()
            """) or ""

            if href_found:
                href = href_found
                break
                
            # Se não achou o link e está na página de busca geral, tenta clicar na aba "Consulta Processual"
            if not clicou_consulta and ("/busca" in url_atual or "q=" in url_atual):
                clicou = await tab.evaluate("""
                    (() => {
                        const els = document.querySelectorAll('a, button, span, div');
                        for (const el of els) {
                            if (el.innerText && el.innerText.trim().toLowerCase() === 'consulta processual') {
                                el.click();
                                return true;
                            }
                        }
                        return false;
                    })()
                """)
                if clicou:
                    clicou_consulta = True
                    log.debug("  Clicou na aba 'Consulta Processual'. Aguardando resultados...")
                    await asyncio.sleep(2)
                    continue

            # Não encontrado?
            if tentativa > 5:
                conteudo = await tab.get_content()
                t = conteudo.lower()
                if any(s in t for s in [
                    "nenhum resultado", "não encontrado", "no results found",
                    "0 resultados", "nenhum processo encontrado",
                    "nenhum documento encontrado"
                ]) and "buscando informações do processo" not in t and "buscando processo" not in t:
                    return "Não encontrado"

            await asyncio.sleep(1)

        if not href:
            return "Não encontrado"

        # 5. Navega para a página do processo
        if href.startswith("/"):
            href = "https://www.jusbrasil.com.br" + href

        url_agora = await _url_atual(tab)
        if url_agora != href:
            log.info(f"  Navegando para: {href}")
            await asyncio.wait_for(tab.get(href), timeout=30)
            await _resolver_cloudflare(tab)
            await asyncio.sleep(3)

        # 6. Valida CNJ na página final e aguarda carregamento completo
        try:
            # Jusbrasil às vezes mostra uma tela de "Buscando processo..." que demora
            tempo_limite_carregamento = time.monotonic() + 45 # Aguarda até 45s para processos antigos
            while time.monotonic() < tempo_limite_carregamento:
                url_agora = await _url_atual(tab)
                
                # Se o Jusbrasil redirecionou dizendo que não encontrou
                if "redirect=lawsuit_not_found" in url_agora or "redirect=not_found" in url_agora:
                    log.warning(f"  Jusbrasil retornou que o processo não foi encontrado.")
                    return "Não encontrado"
                    
                page_text = await tab.evaluate("document.body.innerText") or ""
                page_text_lower = page_text.lower()
                
                # Se ainda estiver na tela de busca animada ou URL intermediária, espera
                if "buscando informações do processo" in page_text_lower or "buscando processo" in page_text_lower or "/processos/consulta/" in url_agora:
                    log.debug("  Aguardando Jusbrasil buscar o processo nos tribunais...")
                    await asyncio.sleep(2)
                    continue
                    
                # Se saiu da tela de busca, valida o CNJ
                page_text_digits = re.sub(r"[^\d]", "", page_text)
                url_digits = re.sub(r"[^\d]", "", url_agora)

                if (cnj not in page_text and
                    cnj_digits not in page_text_digits and
                    cnj_final_13 not in page_text_digits and
                    cnj_final_13 not in url_digits):
                    log.warning(f"  Página não contém CNJ {cnj} — processo errado")
                    return "Não encontrado"
                
                # Se chegou aqui, a página carregou e é o processo certo
                break
                
        except Exception:
            pass

        # 7. Expande seção Histórico e faz scroll para carregar tudo
        try:
            await tab.evaluate("""
                (() => {
                    const els = document.querySelectorAll('button, a, span, h2, h3, summary, div');
                    for (const el of els) {
                        if (el.innerText && el.innerText.trim().toLowerCase() === 'histórico') {
                            el.click();
                        }
                    }
                })()
            """)
            await asyncio.sleep(1)
            
            # Faz scroll para baixo algumas vezes para garantir que o histórico carregue
            for _ in range(4):
                await tab.evaluate("window.scrollBy(0, 800);")
                await asyncio.sleep(1)
                
            # Volta pro topo
            await tab.evaluate("window.scrollTo(0, 0);")
            await asyncio.sleep(1)
            
        except Exception:
            pass

        # 8. Extrai texto do histórico
        texto_historico = await tab.evaluate("""
            (() => {
                // Tenta pegar o texto de toda a seção principal que contém o histórico
                const mainContent = document.querySelector('main') || document.body;
                return mainContent.innerText;
            })()
        """) or ""

        if len(texto_historico) < 80:
            log.debug(f"  Histórico vazio ou curto para {cnj}")
            return "Não encontrado"

        log.debug(f"  Histórico ({len(texto_historico)} chars): {texto_historico[:200]!r}")

        classe = _classificar(texto_historico)
        data   = _extrair_data(texto_historico)

        if classe == "baixado":
            return f"Sim | {data}" if data else "Sim"
        if classe == "publicacao":
            return f"Publicação | {data}" if data else "Publicação"
        return "Não"

    except asyncio.TimeoutError:
        log.warning(f"  Timeout ao pesquisar {cnj}")
        return "Erro"
    except RuntimeError as e:
        if "StopIteration" in str(e):
            log.warning(f"  StopIteration interno ao pesquisar {cnj} — pulando")
        else:
            log.warning(f"  Erro ao pesquisar {cnj}: {e}")
        return "Erro"
    except Exception as e:
        log.warning(f"  Erro ao pesquisar {cnj}: {e}")
        return "Erro"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
async def _run(args):
    caminho_base = Path(args.arquivo)
    if not caminho_base.exists():
        log.error(f"Arquivo não encontrado: {caminho_base}")
        sys.exit(1)

    saida = caminho_base.with_stem(caminho_base.stem + "_verificado")
    if saida.exists():
        try:
            with zipfile.ZipFile(str(saida), "r") as zf:
                if "xl/workbook.xml" not in zf.namelist():
                    raise ValueError("xlsx inválido")
            log.info(f"Retomando: {saida.name}")
        except Exception as e:
            log.warning(f"Arquivo de resultados corrompido ({e}). Recriando...")
            saida.unlink(missing_ok=True)

    if not saida.exists():
        shutil.copy2(str(caminho_base), str(saida))
        log.info(f"Criado: {saida.name}")

    wb = _abrir_planilha(saida)
    ws = wb.active
    todos = _ler_processos(ws)

    pendentes, concluidos = [], 0
    for p in todos:
        val = str(ws.cell(p.linha, COL_RESULTADO).value or "").strip()
        if val and val not in ("Erro", "Não encontrado"):
            # Resultados definitivos (Sim, Não, Publicação) não reprocessar
            concluidos += 1
        else:
            pendentes.append(p)

    log.info(f"Total: {len(todos)} | Já feitos: {concluidos} | Pendentes: {len(pendentes)}")

    if args.inicio > 1:
        pendentes = pendentes[args.inicio - 1:]
        log.info(f"Iniciando no processo {args.inicio} ({len(pendentes)} restantes)")

    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)

    if not pendentes:
        log.info("Nenhum processo pendente.")
        return

    # Perfil persistente: mantém cookies e sessão de login entre execuções
    Path(PERFIL_CHROME).mkdir(exist_ok=True)

    is_ci = bool(os.environ.get("CI"))

    # Mata instâncias zumbi do Chrome usando esse perfil (impede o nodriver
    # de conectar). Usa o lockfile do perfil pra encontrar processos antigos.
    if sys.platform == "win32":
        import subprocess
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "chrome.exe", "/FI", f"WINDOWTITLE ne *"],
                capture_output=True, timeout=5,
            )
        except Exception:
            pass
    # Remove arquivos de lock do perfil
    for lock_name in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
        try:
            (Path(PERFIL_CHROME) / lock_name).unlink(missing_ok=True)
        except Exception:
            pass

    browser_args = ["--disable-session-crashed-bubble"]
    if is_ci:
        browser_args += [
            "--disable-gpu",
            "--disable-dev-shm-usage",
            "--disable-extensions",
            "--disable-setuid-sandbox",
            "--window-size=1920,1080",
        ]

    async def _start_browser():
        """Start browser with retries (CI containers can be slow)."""
        kw = dict(
            headless=is_ci,
            sandbox=False,
            browser_args=browser_args,
        )
        if not is_ci:
            kw["user_data_dir"] = PERFIL_CHROME
        last_err = None
        for attempt in range(3):
            try:
                return await uc.start(**kw)
            except Exception as e:
                last_err = e
                log.warning(f"  Browser start attempt {attempt+1} failed: {e}")
                await asyncio.sleep(2 * (attempt + 1))
        raise last_err

    browser = await _start_browser()

    log.info("Chrome iniciado.")
    log.info("NOTA: Se você não estiver vendo a aba do Jusbrasil, ela pode ter aberto em uma janela diferente ou em segundo plano.")
    log.info("Verifique outras janelas do Chrome abertas.")

    sim = nao = pub = nao_enc = erros = 0

    try:
        try:
            tab = await browser.get("https://www.jusbrasil.com.br/consulta-processual/", new_tab=True)
            await tab.bring_to_front()
            await _resolver_cloudflare(tab)
        except Exception:
            pass

        for i, p in enumerate(pendentes, 1):
            log.info(f"[{i}/{len(pendentes)}] {p.cnj}")
            try:
                resultado = await _pesquisar(browser, tab, p.cnj)
            except Exception as e:
                log.error(f"  Erro fatal na pesquisa: {e}")
                # Tenta recriar browser e tab
                try:
                    await browser.stop()
                except Exception:
                    pass
                log.info("  Recriando browser...")
                browser = await _start_browser()
                try:
                    tab = await browser.get("https://www.jusbrasil.com.br/consulta-processual/", new_tab=True)
                    await tab.bring_to_front()
                    await _resolver_cloudflare(tab)
                    await asyncio.sleep(3)
                except Exception:
                    pass
                resultado = "Erro"

            p.resultado = resultado
            log.info(f"  => {resultado}")

            # Delay aleatório entre processos para evitar detecção de bot
            pausa = random.uniform(3.0, 7.0)
            log.debug(f"  Aguardando {pausa:.1f}s...")
            await asyncio.sleep(pausa)

            if resultado.startswith("Sim"):       sim     += 1
            elif resultado.startswith("Não enc"): nao_enc += 1
            elif resultado.startswith("Pub"):     pub     += 1
            elif resultado == "Erro":             erros   += 1
            else:                                 nao     += 1

            _gravar(ws, p)
            if not _salvar_seguro(wb, saida):
                log.error("  FALHA ao salvar planilha — recarregando...")
                wb = _abrir_planilha(saida)
                ws = wb.active
                _gravar(ws, p)
                _salvar_seguro(wb, saida)

    finally:
        try:
            await browser.stop()
        except Exception:
            pass

    log.info("=" * 50)
    log.info(f"Baixados (Sim):        {sim}")
    log.info(f"Não baixados (Não):    {nao}")
    log.info(f"Publicação:            {pub}")
    log.info(f"Não encontrados:       {nao_enc}")
    log.info(f"Erros:                 {erros}")
    log.info(f"Salvo em: {saida.name}")


def main():
    ap = argparse.ArgumentParser(description="Verifica processos no Jusbrasil (Cloudflare automático)")
    ap.add_argument("--arquivo", default=ARQUIVO_PADRAO)
    ap.add_argument("--debug",  action="store_true")
    ap.add_argument("--inicio", type=int, default=1)
    args = ap.parse_args()

    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    try:
        asyncio.run(_run(args))
    except KeyboardInterrupt:
        log.info("Interrompido pelo usuário.")


if __name__ == "__main__":
    main()
